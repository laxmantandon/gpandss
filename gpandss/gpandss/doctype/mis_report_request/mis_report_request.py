# Copyright (c) 2024, laxman and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils import now_datetime, nowdate
from erpnext.accounts.utils import get_fiscal_year

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

from io import BytesIO
import re

class MISReportRequest(Document):
	def before_save(self):
		if self.items:
			fiscal_year_from = get_fiscal_year(self.from_date)[0]
			fiscal_year_to = get_fiscal_year(self.to_date)[0]

			for report in self.items:
				if report.report_name in ["MIS Balance Sheet","MIS Profit and Loss Statement", "Indirect Cash Flow"]:
					report.filters = {
						"company": self.company,
						"filter_based_on": "Date Range",
						"period_start_date": str(self.from_date),
						"period_end_date": str(self.to_date),
						"from_fiscal_year": fiscal_year_from,
                        "to_fiscal_year": fiscal_year_to,
						"periodicity": "Monthly",
						"cost_center": [],
						"batch":[],
						"department": [self.department] if self.department else [],
						"cf_code":[],
						"cc_code": [self.cc_code] if self.cc_code else [],
						"branch": [self.branch] if self.branch else [],
						"project": [],
						"include_default_book_entries": 1
					}
				if report.report_name in ["MIS Balance Sheet"]:
					report.filters = {
						"company": self.company,
						"filter_based_on": "Date Range",
						"period_start_date": str(self.from_date),
						"period_end_date": str(self.to_date),
						"from_fiscal_year": fiscal_year_from,
						"to_fiscal_year": fiscal_year_to,
						"periodicity": "Monthly",
						"cost_center": [],
						"batch":[],
						"department": [self.department] if self.department else [],
						"cf_code":[],
						"cc_code": [self.cc_code] if self.cc_code else [],
						"branch": [self.branch] if self.branch else [],
						"project": [],
						"include_default_book_entries": 1,
						"accumulated_values": 1,
					}
				if report.report_name == "MIS Trial Balance":
					report.filters = {
						"company": self.company,
						"from_date": str(self.from_date),
						"to_date": str(self.to_date),
						"fiscal_year": fiscal_year_from,
						"batch":[],
						"department": [self.department] if self.department else [],
						"cf_code":[],
						"cc_code": [self.cc_code] if self.cc_code else [],
						"branch": [self.branch] if self.branch else [],
						"with_period_closing_entry": 1,
						"include_default_book_entries": 1,
						"show_net_values": 1
					}
				
				if report.report_name in ["Cashflow Statement Projection Monthwise"]:
					report.filters = {
						"from_date": str(self.from_date),
						"to_date": str(self.to_date),
					}


	def on_submit(self):
		if self.items:
			wb = Workbook()
			for report in self.items:
				x = frappe.call(
					"frappe.desk.query_report.run",
					report_name=report.report_name,
					filters=report.filters,
					ignore_prepared_report=True
				)

				for r in x.get("result"):

					if r.get("is_group") == 0:
						r["color_code"] = "FFFFFF"
						
					else:
						r["font_format"] = "bold"
						if r.get("indent") == 0:
							r["color_code"] = "6699FF"
					
						if r.get("indent") == 1:
							r["color_code"] = "99CCFF" # "33CCFF"
					
						if r.get("indent") == 2:
							r["color_code"] = "CCECFF"
					
						if r.get("indent") == 3:
							r["color_code"] = "CCFFFF"

						if r.get("indent") == 4:
							r["color_code"] = "CCFFCC"


				self.generate_excel(wb, columns=x.get("columns"), data=x.get("result"), sheet_name=report.report_name)

		output_file=f"{self.name}.xlsx"
		file_path = frappe.utils.get_site_path("private", "files", output_file)

		if "Sheet" in wb.sheetnames:
			del wb["Sheet"]

		wb.save(file_path)
		self.attach_file(output_file)


	def generate_excel(self, wb, columns, data, sheet_name):
		columns = frappe.parse_json(columns or [])
		data = frappe.parse_json(data or [])

		ws = wb.create_sheet(sheet_name)

		header = [col.get("label") for col in columns]

		for col_idx, col_name in enumerate(header, start=1):
			cell = ws.cell(row=1, column=col_idx, value=col_name)
			cell.alignment = Alignment(horizontal="center")
			cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
			cell.font = Font(bold=True)


		for idx, col in enumerate(columns, start=1):
			if "width" in col:
				width = int(col["width"]) / 7
				col_letter = get_column_letter(idx)
				ws.column_dimensions[col_letter].width = width


		for row_idx, row in enumerate(data, start=2):
			for col_idx, col in enumerate(columns, start=1):
				cell_value = row.get(col["fieldname"], "")
				cell = ws.cell(row=row_idx, column=col_idx)

				cell_str = str(cell_value).strip()
				is_numeric = False

				try:
					numeric_val = float(cell_value)
					cell.value = numeric_val
					cell.number_format = '#,##0.00'
					is_numeric = True
				except (ValueError, TypeError):
					cell.value = cell_str

				# Alignment logic for data rows
				if any(char.isalpha() for char in cell_str):
					cell.alignment = Alignment(horizontal="left")
				else:
					cell.alignment = Alignment(horizontal="right")

				color_code = row.get("color_code", 'FFFFFF')
				cell.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")

				if row.get("font_format"):
					cell.font = Font(bold=True)



	def attach_file(self, output_file):
		frappe.get_doc(
			{
				"doctype": "File",
				"attached_to_doctype": self.doctype,
				"attached_to_name": self.name,
				"file_name": output_file,
				"file_url": f"/private/files/{output_file}",
				"is_private": 1
			}
		).save()