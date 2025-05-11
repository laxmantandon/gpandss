import frappe
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from io import BytesIO


# def export_to_excel(columns, data, output_file="output.xlsx"):
@frappe.whitelist()
def export_to_excel(columns=None, data=None, report_name="Output"):
    output_file=f"{report_name}.xlsx"
    
    columns = frappe.parse_json(columns or [])
    data = frappe.parse_json(data or [])

    wb = Workbook()
    ws = wb.active

    header = [col.get("label") for col in columns]

    for col_idx, col_name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color=col_name.get("color_code"), end_color=col_name.get("color_code"), fill_type="solid")

    for idx, col in enumerate(columns, start=1):
        if "width" in col:
            width = int(col["width"]) / 7  # Approximation for Excel width
            ws.column_dimensions[chr(64 + idx)].width = width


    for row_idx, row in enumerate(data, start=2):
        for col_idx, col in enumerate(columns, start=1):
            cell_value = row.get(col["fieldname"], "")
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)

            color_code = row.get("color_code", 'FFFFFF')
            cell.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")


    file_path = frappe.utils.get_site_path("public", "files", output_file)
    wb.save(file_path)

    return f'/files/{output_file}'

    # frappe.response["filename"] = output_file
    # frappe.response["filecontent"] = open(file_path, "rb").read()
    # frappe.response["type"] = "binary"




# # Example usage

# export_to_excel(columns, data)
