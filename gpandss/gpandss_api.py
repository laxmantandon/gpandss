from datetime import datetime
import json
import random
import frappe
import base64
import os
from frappe.utils import get_files_path, get_site_name, now
import requests
from frappe.utils.password import check_password, get_password_reset_limit


@frappe.whitelist(allow_guest=True)
def receive_message():
    try:
        message = frappe.form_dict
        frappe.publish_realtime(event='new_chat_message', user= message.get('receiverId'), message={'user': message.get('receiverId'), 'message': message})
        frappe.response.message={
            'status':True,
            'messgae':'inserted'
        }

    except Exception as e:
        frappe.response.message={
            'status':False,
            'messgae':f"{e}"
        }

    
@frappe.whitelist()
def upload_file_in_doctype(datas, filename, docname, doctype):
   for data in datas:
        try:
            filename_ext = f'/home/frappe/frappe-bench/sites/ss.erpdesks.com/private/files/{filename}.png'
            base64data = data.replace('data:image/jpeg;base64,', '')
            imgdata = base64.b64decode(base64data)
            with open(filename_ext, 'wb') as file:
                file.write(imgdata)

            doc = frappe.get_doc(
                {
                    "file_name": f'{filename}.png',
                    "is_private": 1,
                    "file_url": f'/private/files/{filename}.png',
                    "attached_to_doctype": doctype if doctype else "Geo Mitra Ledger Report",
                    "attached_to_name": docname,
                    "doctype": "File",
                }
            )
            doc.flags.ignore_permissions = True
            doc.insert()
            frappe.db.commit()
            return doc.file_url

        except Exception as e:
            frappe.log_error('ng_write_file', str(e))
            return e



@frappe.whitelist()
def upload_file_document_in_doctype(datas, filename, docname, doctype):
   for data in datas:
        try:
            file_ext = data.get('ext')
            base64data = data.get('base64')

            if not file_ext or not base64data:
                frappe.throw("Invalid file data. Missing extension or base64 content.")

            if "," in base64data:
                base64data = base64data.split(",")[-1]

            # Define file path
            file_path = f'/home/frappe/frappe-bench/sites/ss.erpdesks.com/private/files/{filename}.{file_ext}'

            # Decode and save file
            file_data = base64.b64decode(base64data)
            with open(file_path, 'wb') as file:
                file.write(file_data)

            # Create File document in Frappe
            doc = frappe.get_doc({
                "file_name": f"{filename}.{file_ext}",
                "is_private": 1,
                "file_url": f"/private/files/{filename}.{file_ext}",
                "attached_to_doctype": doctype if doctype else "Geo Mitra Ledger Report",
                "attached_to_name": docname,
                "doctype": "File",
            })
            doc.flags.ignore_permissions = True
            doc.insert()
            frappe.db.commit()

            return doc.file_url

        except Exception as e:
            frappe.log_error('upload_file_document_in_doctype', str(e))
            return str(e)


@frappe.whitelist()
def get_doctype_images(doctype, docname, is_private):
    attachments = frappe.db.get_all("File",
        fields=["attached_to_name", "file_name", "file_url", "is_private"],
        filters={"attached_to_name": docname, "attached_to_doctype": doctype}
    )
    resp = []
    for attachment in attachments:
        # file_path = site_path + attachment["file_url"]
        x = get_files_path(attachment['file_name'], is_private=is_private)
        with open(x, "rb") as f:
            # encoded_string = base64.b64encode(image_file.read())
            img_content = f.read()
            img_base64 = base64.b64encode(img_content).decode()
            img_base64 = 'data:image/jpeg;base64,' + img_base64
        resp.append({"image": img_base64})

    return resp

@frappe.whitelist()
def generate_keys(user):
    user_details = frappe.get_doc("User", user)
    api_secret = frappe.generate_hash(length=15)
    
    if not user_details.api_key:
        api_key = frappe.generate_hash(length=15)
        user_details.api_key = api_key
    
    user_details.api_secret = api_secret

    user_details.flags.ignore_permissions = True
    user_details.save(ignore_permissions = True)
    frappe.db.commit()
    
    return user_details.api_key, api_secret


@frappe.whitelist(allow_guest=True)
def login_user(usr, pwd):

    if not usr or not pwd:
        frappe.local.response["message"] = {
            "status": False,
            "message": "invalid inputs"
        }
        return
    user_email = ""
    user_exist = frappe.db.count("User",{'email': usr})
    if user_exist > 0:
        userm = frappe.db.get_all('User', filters={'email': usr}, fields=['*'])
        user_email = userm[0].name
        try:
            check_password(user_email, pwd)
        except Exception as e:
            frappe.local.response["message"] = {
                "status": False,
                "message": "User Password  Is Not Correct",
            }
            return



        api_key, api_secret = generate_keys(user_email)
        # frappe.local.login_manager.user = user_email
        # frappe.local.login_manager.post_login()
        geomitra_data = frappe.db.get_all('Employee', filters={'user_id': user_email}, fields=['*'])
        if geomitra_data :
            frappe.local.response["message"] = {
                "status": True,
                "message": "User Already Exists",
                "data":{
                "api_key": api_key,
                "api_secret": api_secret,
                "first_name": userm[0].first_name
                }
            }
            return        

    frappe.local.response["message"] = {
        "status": False,
        "message": "User Not Exists",
    }



@frappe.whitelist()
def export_financial_reports(output_filename="Financial_Reports.xlsx"):
   
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Alignment
    import frappe

    wb = Workbook()

    financial_reports = [
        {"report_name": "Profit and Loss Statement", "method": "erpnext.accounts.report.profit_and_loss_statement.profit_and_loss_statement"},
        {"report_name": "General Ledger", "method": "erpnext.accounts.report.general_ledger.general_ledger"},
        {"report_name": "Balance Sheet", "method": "erpnext.accounts.report.balance_sheet.balance_sheet"}
    ]

    # Fetch and add each report to a new sheet
    for index, report in enumerate(financial_reports):
        sheet_name = report["report_name"]
        
        # Fetch report data
        report_data = frappe.get_attr(report["method"])(filters={})
        columns = report_data.get("columns", [])
        data = report_data.get("result", [])

        # Create a new sheet
        if index == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        # Apply header style
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
        data_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # Light blue

        # Write headers
        header = [col.get("label", col.get("fieldname")) for col in columns]
        for col_idx, col_name in enumerate(header, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = header_fill

        # Set column widths
        for idx, col in enumerate(columns, start=1):
            if "width" in col:
                width = int(col["width"]) / 7  # Approximation for Excel width
                ws.column_dimensions[chr(64 + idx)].width = width

        # Write data rows
        for row_idx, row in enumerate(data, start=2):
            for col_idx, col in enumerate(columns, start=1):
                fieldname = col.get("fieldname", "")
                cell_value = row.get(fieldname, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                color_code = row.get("color_code", 'FFFFFF')
                cell.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")

    # Save workbook
    file_path = frappe.utils.get_site_path("public", "files", output_filename)
    wb.save(file_path)

    return f'/files/{output_filename}'
